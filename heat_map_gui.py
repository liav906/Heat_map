import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
from tkinter import Tk, Label, Entry, Button, filedialog, messagebox, PhotoImage
from tkcalendar import DateEntry
from tkinter import ttk

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    file_entry.delete(0, 'end')
    file_entry.insert(0, file_path)

def process_file():
    start_date = start_date_entry.get_date()
    end_date = end_date_entry.get_date()
    file_path = file_entry.get()

    if not file_path:
        messagebox.showwarning("Input Error", "Please select an Excel file.")
        return

    if start_date == end_date:
        output_file_name = f"{start_date.strftime('%d.%m.%y')}_heat_map"
    else:
        output_file_name = f"{start_date.strftime('%d')}-{end_date.strftime('%d.%m.%y')}_heat_map"

    output_folder = os.path.join(os.path.expanduser('~'), 'Desktop', output_file_name)
    os.makedirs(output_folder, exist_ok=True)

    df = pd.read_excel(file_path)

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    device_counter = 0
    gps_error_cnt = 0
    clock_error = 0
    total_error = 0
    good_reports = 0

    previous_times = {}
    si_error_counts = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "SI Data"

    col_offset = 1
    for si in df['si'].unique():
        si_data = df[df['si'] == si][['lat', 'created_at']].copy()
        si_data['created_at'] = pd.to_datetime(si_data['created_at'], format='%Y-%m-%d %H:%M:%S')

        ws.cell(row=1, column=col_offset).value = si
        ws.cell(row=2, column=col_offset).value = "created_at"

        previous_times[si] = None

        row_offset = 3
        for i, row in si_data.iterrows():
            lat = row['lat']
            created_at = row['created_at']

            if si not in si_error_counts:
                si_error_counts[si] = {'total': 0, 'gps_error_cnt': 0, 'clock_error': 0, 'total_error': 0}
                device_counter += 1

            si_error_counts[si]['total'] += 1

            row_fill = green_fill
            if lat == "error":
                gps_error_cnt += 1
                si_error_counts[si]['gps_error_cnt'] += 1
                row_fill = yellow_fill
            else:
                good_reports += 1

            if previous_times[si]:
                time_diff = (created_at - previous_times[si]).total_seconds()
                if time_diff > 1800:
                    clock_error += 1
                    si_error_counts[si]['clock_error'] += 1
                    if row_fill == yellow_fill:
                        row_fill = red_fill
                        total_error += 1
                        si_error_counts[si]['total_error'] += 1
                        gps_error_cnt -= 1
                        si_error_counts[si]['gps_error_cnt'] -= 1
                        clock_error -= 1
                        si_error_counts[si]['clock_error'] -= 1
                    else:
                        row_fill = orange_fill

            ws.cell(row=row_offset, column=col_offset).value = created_at
            ws.cell(row=row_offset, column=col_offset).fill = row_fill

            previous_times[si] = created_at
            row_offset += 1

        col_offset += 1

    modified_file_path = os.path.join(output_folder, 'modified_' + output_file_name + '.xlsx')
    wb.save(modified_file_path)

    def generate_pie_chart(sizes, labels, colors, title, filename):
        plt.figure()
        plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%',
                shadow=True, startangle=140)
        plt.axis('equal')
        plt.title(title)
        plt.savefig(os.path.join(output_folder, filename))
        plt.close()

    total_labels = []
    total_sizes = []
    total_colors = []

    if gps_error_cnt:
        total_labels.append('GPS Errors')
        total_sizes.append(gps_error_cnt)
        total_colors.append('yellow')

    if clock_error:
        total_labels.append('Clock Errors')
        total_sizes.append(clock_error)
        total_colors.append('orange')

    if total_error:
        total_labels.append('Total Errors')
        total_sizes.append(total_error)
        total_colors.append('red')

    if good_reports:
        total_labels.append('Good Reports')
        total_sizes.append(good_reports)
        total_colors.append('green')

    generate_pie_chart(total_sizes, total_labels, total_colors, 'Total Errors', 'total_error_pie_chart.png')

    for si, counts in si_error_counts.items():
        labels = []
        sizes = []
        colors = []

        if counts['gps_error_cnt']:
            labels.append('GPS Errors')
            sizes.append(counts['gps_error_cnt'])
            colors.append('yellow')

        if counts['clock_error']:
            labels.append('Clock Errors')
            sizes.append(counts['clock_error'])
            colors.append('orange')

        if counts['total_error']:
            labels.append('Total Errors')
            sizes.append(counts['total_error'])
            colors.append('red')

        good_count = counts['total'] - counts['gps_error_cnt'] - counts['clock_error'] - counts['total_error']
        if good_count:
            labels.append('Good Reports')
            sizes.append(good_count)
            colors.append('green')

        generate_pie_chart(sizes, labels, colors, f'Errors for SI {si}', f'{si}_error_pie_chart.png')

    messagebox.showinfo("Job Done", "The processing is complete!")

# Create the main window
root = Tk()
root.title("Excel Processing Tool")

# # Load the logo image
# logo_path = "/Users/liav/PycharmProjects/time_color/icon.png"
# if not os.path.exists(logo_path):
#     messagebox.showerror("File Not Found", f"The logo file was not found at: {logo_path}")
# else:
#     logo_img = PhotoImage(file=logo_path)
#     logo_label = Label(root, image=logo_img)
#     logo_label.grid(row=0, column=0, columnspan=3, pady=10)

# Create GUI elements
Label(root, text="Start Date:").grid(row=1, column=0, padx=10, pady=10)
start_date_entry = DateEntry(root, date_pattern='dd.MM.yyyy')
start_date_entry.grid(row=1, column=1, padx=10, pady=10)

Label(root, text="End Date:").grid(row=2, column=0, padx=10, pady=10)
end_date_entry = DateEntry(root, date_pattern='dd.MM.yyyy')
end_date_entry.grid(row=2, column=1, padx=10, pady=10)

Label(root, text="Select Excel File:").grid(row=3, column=0, padx=10, pady=10)
file_entry = Entry(root)
file_entry.grid(row=3, column=1, padx=10, pady=10)
Button(root, text="Browse", command=select_file).grid(row=3, column=2, padx=10, pady=10)

Button(root, text="Process", command=process_file).grid(row=4, column=0, columnspan=3, padx=10, pady=10)

# Apply a theme using ttk.Style
style = ttk.Style(root)
style.theme_use('clam')

# Customize colors
style.configure('TButton', background='blue', foreground='white')
style.configure('TLabel', background='white', foreground='black')
style.configure('TEntry', background='white', foreground='black')

# Run the GUI
root.mainloop()