
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt

# Load the Excel file
file_path = 'modified_test.xlsx'
df = pd.read_excel("/Users/liav/PycharmProjects/time_color/modified_test.xlsx")


# Initialize fill patterns
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Initialize counters
device_counter = 0
gps_error_cnt = 0
clock_error = 0
total_error = 0
good_reports = 0

# Initialize a dictionary to track the previous timestamp for each si and error counts per si
previous_times = {}
si_error_counts = {}

# Create a new workbook and a single worksheet
wb = Workbook()
ws = wb.active
ws.title = "SI Data"

# Process data and populate the worksheet
col_offset = 1
for si in df['si'].unique():
    si_data = df[df['si'] == si][['lat', 'created_at']].copy()
    si_data['created_at'] = pd.to_datetime(si_data['created_at'], format='%Y-%m-%d %H:%M:%S')

    # Add si as header in the first row
    ws.cell(row=1, column=col_offset).value = si

    # Add title "created_at" in the second row
    ws.cell(row=2, column=col_offset).value = "created_at"

    # Initialize the previous time for the current 'si'
    previous_times[si] = None

    # Process each row in the si_data
    row_offset = 3
    for i, row in si_data.iterrows():
        lat = row['lat']
        created_at = row['created_at']

        if si not in si_error_counts:
            si_error_counts[si] = {'total': 0, 'gps_error_cnt': 0, 'clock_error': 0, 'total_error': 0}
            device_counter += 1

        # Increment the total count for the current si
        si_error_counts[si]['total'] += 1

        # Determine the fill color
        row_fill = green_fill  # Default to green
        if lat == "error":
            gps_error_cnt += 1
            si_error_counts[si]['gps_error_cnt'] += 1
            row_fill = yellow_fill
        else:
            good_reports += 1

        if previous_times[si]:
            time_diff = (created_at - previous_times[si]).total_seconds()
            if time_diff > 1800:
                clock_error += 1
                si_error_counts[si]['clock_error'] += 1
                if row_fill == yellow_fill:
                    row_fill = red_fill
                    total_error += 1
                    si_error_counts[si]['total_error'] += 1
                    gps_error_cnt -= 1
                    si_error_counts[si]['gps_error_cnt'] -= 1
                    clock_error -= 1
                    si_error_counts[si]['clock_error'] -= 1
                else:
                    row_fill = orange_fill

        # Add the created_at value to the worksheet
        ws.cell(row=row_offset, column=col_offset).value = created_at

        # Apply color to the cell
        ws.cell(row=row_offset, column=col_offset).fill = row_fill

        # Update previous time for the current 'si'
        previous_times[si] = created_at
        row_offset += 1

    # Move to the next column for the next 'si'
    col_offset += 1

# Save the modified Excel file
modified_file_path = 'modified_' + file_path
wb.save(modified_file_path)

# Function to generate pie chart
def generate_pie_chart(sizes, labels, colors, title, filename):
    plt.figure()
    plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%',
            shadow=True, startangle=140)
    plt.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
    plt.title(title)
    plt.savefig(filename)
    plt.close()

# Generate the total pie chart
total_labels = []
total_sizes = []
total_colors = []

if gps_error_cnt:
    total_labels.append('GPS Errors')
    total_sizes.append(gps_error_cnt)
    total_colors.append('yellow')

if clock_error:
    total_labels.append('Clock Errors')
    total_sizes.append(clock_error)
    total_colors.append('orange')

if total_error:
    total_labels.append('Total Errors')
    total_sizes.append(total_error)
    total_colors.append('red')

if good_reports:
    total_labels.append('Good Reports')
    total_sizes.append(good_reports)
    total_colors.append('green')

generate_pie_chart(total_sizes, total_labels, total_colors, 'Total Errors', 'total_error_pie_chart.png')

# Generate pie charts for each si
for si, counts in si_error_counts.items():
    labels = []
    sizes = []
    colors = []

    if counts['gps_error_cnt']:
        labels.append('GPS Errors')
        sizes.append(counts['gps_error_cnt'])
        colors.append('yellow')

    if counts['clock_error']:
        labels.append('Clock Errors')
        sizes.append(counts['clock_error'])
        colors.append('orange')

    if counts['total_error']:
        labels.append('Total Errors')
        sizes.append(counts['total_error'])
        colors.append('red')

    good_count = counts['total'] - counts['gps_error_cnt'] - counts['clock_error'] - counts['total_error']
    if good_count:
        labels.append('Good Reports')
        sizes.append(good_count)
        colors.append('green')

    generate_pie_chart(sizes, labels, colors, f'Errors for SI {si}', f'{si}_error_pie_chart.png')

# Print device counter and errors
print(f'Total devices: {device_counter}')
print(f'GPS errors: {gps_error_cnt}')
print(f'Clock errors: {clock_error}')
print(f'Total errors: {total_error}')
print(f'Good reports: {good_reports}')